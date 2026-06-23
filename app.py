from dotenv import load_dotenv
load_dotenv()

from flask import Flask, Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from functools import wraps
import os
import functools
import random
import sqlite3
import time
import io
import threading
from datetime import datetime, timedelta
from werkzeug.security import check_password_hash, generate_password_hash

from services import ycloud, wisphub, genieacs

# ---------------------------------------------------------------------------
# App y Blueprints
# ---------------------------------------------------------------------------

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY") or os.urandom(24)
app.permanent_session_lifetime = timedelta(hours=8)
app.config['SESSION_REFRESH_EACH_REQUEST'] = True
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax'
)

# Diccionario para rate limiting por IP (Fuerza Bruta)
ip_failed_attempts = {}

def get_ip_block_time(ip):
    now = time.time()
    record = ip_failed_attempts.get(ip)
    if record:
        if record['blocked_until'] and now < record['blocked_until']:
            return int(record['blocked_until'] - now)
        elif record['blocked_until']:
            del ip_failed_attempts[ip]
    return 0

def record_failed_attempt(ip):
    now = time.time()
    record = ip_failed_attempts.get(ip, {'attempts': 0, 'blocked_until': None})
    if record['blocked_until'] and now >= record['blocked_until']:
        record = {'attempts': 0, 'blocked_until': None}
    record['attempts'] += 1
    if record['attempts'] >= 5:
        record['blocked_until'] = now + 300 # Bloqueo por 5 minutos
    ip_failed_attempts[ip] = record

def clear_failed_attempts(ip):
    if ip in ip_failed_attempts:
        del ip_failed_attempts[ip]

# Blueprint admin (prefijo /admin, nombre 'admin' para que url_for('admin.XXX') funcione)
admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

@admin_bp.context_processor
def inject_empresa():
    return {'empresa': os.getenv('EMPRESA', 'SkyPass')}

# ---------------------------------------------------------------------------
# Base de datos
# ---------------------------------------------------------------------------

def get_db_connection():
    conn = sqlite3.connect('skypass.db')
    conn.row_factory = sqlite3.Row
    return conn



# ===========================================================================
# ██╗   ██╗███████╗██╗   ██╗ █████╗ ██████╗ ██╗ ██████╗
# ██║   ██║██╔════╝██║   ██║██╔══██╗██╔══██╗██║██╔═══██╗
# ██║   ██║███████╗██║   ██║███████║██████╔╝██║██║   ██║
# ██║   ██║╚════██║██║   ██║██╔══██║██╔══██╗██║██║   ██║
# ╚██████╔╝███████║╚██████╔╝██║  ██║██║  ██║██║╚██████╔╝
#  ╚═════╝ ╚══════╝ ╚═════╝ ╚═╝  ╚═╝╚═╝  ╚═╝╚═╝ ╚═════╝
# ===========================================================================


# ---------------------------------------------------------------------------
# Límites de cambios mensuales (usuario)
# ---------------------------------------------------------------------------

def obtener_cambios_por_mes_global():
    try:
        conn = get_db_connection()
        cur = conn.execute("SELECT valor FROM admin_settings WHERE clave = 'max_cambios_mes'")
        data = cur.fetchone()
        conn.close()
        if data and data['valor']:
            return int(data['valor'])
    except Exception as e:
        print(f"Error obteniendo max_cambios_mes: {e}")
    return 2


def contar_cambios_usuario_mes(cedula):
    ahora = datetime.now()
    inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    try:
        conn = get_db_connection()
        cur = conn.execute(
            "SELECT COUNT(*) as total FROM change_history WHERE cedula = ? AND fecha >= ?",
            (cedula, inicio_mes.isoformat())
        )
        data = cur.fetchone()
        conn.close()
        return data['total'] if data else 0
    except Exception as e:
        print(f"Error contando cambios del usuario: {e}")
        return 0


def obtener_limite_cliente(ip):
    try:
        conn = get_db_connection()
        cur = conn.execute("SELECT limite_personalizado FROM user_limits WHERE ip = ?", (ip,))
        data = cur.fetchone()
        conn.close()
        if data and data['limite_personalizado']:
            return int(data['limite_personalizado'])
    except Exception as e:
        print(f"Error obteniendo límite personalizado: {e}")
    return obtener_cambios_por_mes_global()


def limpiar_historial_antiguo():
    ahora = datetime.now()
    primer_dia_mes_actual = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if primer_dia_mes_actual.month == 1:
        primer_dia_mes_anterior = primer_dia_mes_actual.replace(year=primer_dia_mes_actual.year - 1, month=12)
    else:
        primer_dia_mes_anterior = primer_dia_mes_actual.replace(month=primer_dia_mes_actual.month - 1)
    try:
        conn = get_db_connection()
        conn.execute("DELETE FROM change_history WHERE fecha < ?", (primer_dia_mes_anterior.isoformat(),))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error limpiando historial antiguo: {e}")


def registrar_cambio_usuario(cedula, tipo_cambio, valor_nuevo):
    try:
        conn = get_db_connection()
        conn.execute(
            "INSERT INTO change_history (admin_id, cedula, tipo_cambio, valor_nuevo, fecha) VALUES (?, ?, ?, ?, ?)",
            (None, cedula, tipo_cambio, valor_nuevo, datetime.now().isoformat())
        )
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error al registrar cambio (usuario): {e}")
        return False


# ---------------------------------------------------------------------------
# Sesión del cliente
# ---------------------------------------------------------------------------

def cliente_requerido(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("cliente_encontrado"):
            return redirect(url_for("index"))
        return func(*args, **kwargs)
    return wrapper


def obtener_cliente_actual():
    return {
        'nombre':     session.get('nombre', ''),
        'cedula':     session.get('cedula', ''),
        'celular':    session.get('telefono', ''),
        'poblacion':  session.get('poblacion', ''),
        'plan_megas': session.get('plan_megas', '')
    }


# ---------------------------------------------------------------------------
# Filtros Jinja
# ---------------------------------------------------------------------------

def formatear_fecha(value, formato='%d/%m/%Y %H:%M'):
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value.replace('Z', '+00:00'))
        except Exception:
            return value
    return value.strftime(formato)

app.jinja_env.filters['formatear_fecha'] = formatear_fecha


# ---------------------------------------------------------------------------
# Rutas — cliente
# ---------------------------------------------------------------------------

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "GET":
        session.clear()
        return render_template("users/user_login.html",
                               empresa=os.getenv('EMPRESA', 'SkyPass'),
                               soporte_wa=os.getenv('SOPORTE_WHATSAPP', ''))

    # Obtener IP real del cliente (soporta proxies como Nginx/Cloudflare)
    ip = request.headers.get('X-Forwarded-For', request.remote_addr)
    if ip and ',' in ip:
        ip = ip.split(',')[0].strip()
    
    block_time = get_ip_block_time(ip)
    if block_time > 0:
        return jsonify({'success': False, 'code': 'bloqueo_ip',
                        'message': f'Tu IP ha sido bloqueada temporalmente por múltiples intentos fallidos. Intenta en {block_time // 60} min y {block_time % 60} seg.'})

    intentos_cedula = session.get('intentos_cedula', 0)
    bloqueo_cedula_hasta = session.get('bloqueo_cedula_hasta')
    ahora = datetime.now()

    if bloqueo_cedula_hasta:
        try:
            bloqueo_dt = datetime.fromisoformat(bloqueo_cedula_hasta)
            if ahora < bloqueo_dt:
                seg = int((bloqueo_dt - ahora).total_seconds())
                return jsonify({'success': False, 'code': 'bloqueo_cedula',
                                'message': f'Has superado el número de intentos. Intenta de nuevo en {seg // 60} minutos y {seg % 60} segundos.'})
        except Exception:
            pass

    cedula = request.form.get("cedula")
    if not cedula:
        record_failed_attempt(ip)
        intentos_cedula += 1
        session['intentos_cedula'] = intentos_cedula
        restantes = 3 - intentos_cedula
        if intentos_cedula >= 3:
            session['bloqueo_cedula_hasta'] = (ahora + timedelta(minutes=2)).isoformat()
            session['intentos_cedula'] = 0
            return jsonify({'success': False, 'code': 'bloqueo_cedula',
                            'message': 'Has superado el número de intentos permitidos. Tu acceso ha sido bloqueado por 2 minutos.'})
        return jsonify({'success': False, 'code': 'sin_cedula',
                        'message': f'Por favor ingrese su cédula. Te quedan {restantes} intento(s) antes de ser bloqueado.'})

    cliente, error = wisphub.buscar_cliente(cedula)
    if error:
        record_failed_attempt(ip)
        intentos_cedula += 1
        session['intentos_cedula'] = intentos_cedula
        restantes = 3 - intentos_cedula
        if intentos_cedula >= 3:
            session['bloqueo_cedula_hasta'] = (ahora + timedelta(minutes=2)).isoformat()
            session['intentos_cedula'] = 0
            return jsonify({'success': False, 'code': 'bloqueo_cedula',
                            'message': 'Has superado el número de intentos permitidos. Tu acceso ha sido bloqueado por 2 minutos.'})
        return jsonify({'success': False, 'code': error['code'],
                        'message': f"{error['message']} Te quedan {restantes} intento(s) antes de ser bloqueado."})

    session.permanent = True
    session.pop('intentos_cedula', None)
    session.pop('bloqueo_cedula_hasta', None)

    # Verificar que el equipo esté registrado en ACS antes de enviar el código
    ip_cliente = cliente.get('ip') or cliente.get('ip_address', '')
    if ip_cliente:
        _, device_id = genieacs.obtener_estado_online_device(ip_cliente)
        if not device_id:
            return jsonify({'success': False, 'code': 'sin_acs',
                            'nombre': cliente.get('nombre', ''),
                            'cedula': cedula})

    whatsapp = ycloud.normalizar_numero(cliente.get('telefono', ''))
    if not whatsapp:
        return jsonify({'success': False, 'code': 'error_numero',
                        'message': 'El número de teléfono no es válido para WhatsApp Colombia.'})

    codigo = str(random.randint(100000, 999999))
    session['codigo_verificacion'] = codigo
    session['telefono'] = whatsapp
    session['nombre'] = cliente.get('nombre', '')
    session['cedula'] = cedula

    if not ycloud.enviar_otp(whatsapp, codigo):
        return jsonify({'success': False, 'code': 'error_envio',
                        'message': 'No se pudo enviar el código por WhatsApp.'})

    return jsonify({'success': True, 'ultimos4': whatsapp[-4:]})


@app.route('/verificar_codigo_ajax', methods=['POST'])
def verificar_codigo_ajax():
    ip = request.headers.get('X-Forwarded-For', request.remote_addr)
    if ip and ',' in ip:
        ip = ip.split(',')[0].strip()
        
    block_time = get_ip_block_time(ip)
    if block_time > 0:
        return jsonify({'success': False, 'code': 'bloqueo_ip',
                        'message': f'Tu IP ha sido bloqueada temporalmente por múltiples intentos fallidos. Intenta en {block_time // 60} min y {block_time % 60} seg.'})

    codigo = request.form.get('codigo')
    intentos = session.get('intentos_codigo', 0)
    bloqueo_hasta = session.get('bloqueo_codigo_hasta')
    ahora = datetime.now()

    if bloqueo_hasta:
        try:
            bloqueo_dt = datetime.fromisoformat(bloqueo_hasta)
            if ahora < bloqueo_dt:
                seg = int((bloqueo_dt - ahora).total_seconds())
                return jsonify({'success': False,
                                'message': f'Has superado el número de intentos. Intenta de nuevo en {seg // 60} minutos y {seg % 60} segundos.'})
        except Exception:
            pass

    if codigo != session.get('codigo_verificacion'):
        record_failed_attempt(ip)
        intentos += 1
        session['intentos_codigo'] = intentos
        restantes = 3 - intentos
        if intentos >= 3:
            session['bloqueo_codigo_hasta'] = (ahora + timedelta(minutes=2)).isoformat()
            session['intentos_codigo'] = 0
            return jsonify({'success': False,
                            'message': 'Has superado el número de intentos permitidos. Tu acceso ha sido bloqueado por 2 minutos.'})
        return jsonify({'success': False,
                        'message': f'Código incorrecto. Te quedan {restantes} intento(s) antes de ser bloqueado.'})

    # Código correcto
    clear_failed_attempts(ip)
    session.pop('intentos_codigo', None)
    session.pop('bloqueo_codigo_hasta', None)
    session.pop('_flashes', None)

    cedula = session.get('cedula')
    _, servicios_unificados = wisphub.obtener_servicios_por_cedula(cedula)

    if len(servicios_unificados) > 1:
        lista = []
        for idx, s in enumerate(servicios_unificados, 1):
            texto = f"Servicio de Internet {idx}"
            if s.get('direccion'):
                texto += f" ({s['direccion']})"
            if s.get('ssid_router_wifi'):
                texto += f" - SSID: {s['ssid_router_wifi']}"
            lista.append({'ip': s.get('ip', ''), 'texto': texto})
        return jsonify({'success': True, 'multiple_servicios': True, 'servicios': lista})

    ip = servicios_unificados[0]['ip'] if servicios_unificados else session.get('ip')
    session['ip'] = ip
    session['cliente_encontrado'] = True
    flash('Sesión iniciada correctamente', 'success')
    return jsonify({'success': True, 'redirect': url_for('dashboard')})


@app.route("/seleccionar_servicio", methods=["POST"])
def seleccionar_servicio():
    ip = request.form.get("ip")
    if not ip:
        return jsonify({'success': False})
    session['ip'] = ip
    session['cliente_encontrado'] = True
    return jsonify({'success': True, 'redirect': url_for('dashboard')})


@app.route("/dashboard")
@cliente_requerido
def dashboard():
    limpiar_historial_antiguo()
    cliente = obtener_cliente_actual()
    ip_seleccionada = session.get('ip')
    estado_servicio = None
    ssid_actual = ''
    password_actual = ''

    # 1. Obtener estado del servicio desde Wisphub
    _, servicios_data = wisphub.obtener_servicios_por_cedula(cliente['cedula'], timeout=7)
    if servicios_data:
        svc = next((s for s in servicios_data if s.get('ip') == ip_seleccionada), servicios_data[0])
        estado_servicio = svc.get('estado', '').lower()

    # 2. Obtener SSID y contraseña reales desde GenieACS
    try:
        device = genieacs._get_svc().get_device_by_ip(ip_seleccionada)
        if device:
            # SSID: primera interfaz WiFi activa
            ssid_actual = genieacs._get_svc().get_ssid_primary(device) or ''
            # Contraseña: VirtualParameters.WlanPassword
            password_actual = genieacs._get_svc()._get_param(device, "VirtualParameters.WlanPassword") or ''
            if password_actual:
                password_actual = str(password_actual)
    except Exception as e:
        print(f"[Dashboard] Error obteniendo WiFi de GenieACS: {e}")

    cambios_realizados = contar_cambios_usuario_mes(cliente['cedula'])
    return render_template(
        "users/user_dashboard.html",
        cliente=cliente,
        cambios_realizados=cambios_realizados,
        estado_servicio=estado_servicio,
        ssid_actual=ssid_actual,
        password_actual=password_actual
    )


@app.route("/cambiar_clave", methods=["GET", "POST"])
@cliente_requerido
def cambiar_clave():
    limpiar_historial_antiguo()
    cliente = obtener_cliente_actual()

    if request.method == "GET":
        return render_template("users/user_cambiar_clave.html", cliente=cliente)

    accion = request.args.get('accion')

    if accion == 'validar_dispositivo':
        ip = session.get('ip')
        if not ip:
            return jsonify({'success': False, 'message': 'No se encontró la IP del cliente'})
        try:
            dispositivo_online, device_id = genieacs.obtener_estado_online_device(ip)
            if not dispositivo_online:
                if not genieacs.verificar_dispositivo_online_alternativo(ip):
                    return jsonify({'success': False, 'message': 'El dispositivo está desconectado'})
                device_id = genieacs.obtener_device_id_por_ip(ip)
                if not device_id:
                    return jsonify({'success': False, 'message': 'No se encontró el dispositivo'})
            interfaces_activas = genieacs.detectar_interfaces_wifi_activas(device_id)
            if not interfaces_activas:
                return jsonify({'success': False, 'message': 'No se encontraron interfaces WiFi activas'})
            session['device_info'] = {
                'device_id': device_id,
                'interfaces_activas': interfaces_activas,
                'detected_at': datetime.now().isoformat()
            }
            return jsonify({'success': True, 'device_id': device_id,
                            'interfaces_activas': interfaces_activas,
                            'interfaces_count': len(interfaces_activas)})
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error validando dispositivo: {e}'})

    if not request.is_json:
        return jsonify({'success': False, 'message': 'Solo se permite el flujo AJAX.'})

    limite_cambios = obtener_limite_cliente(session.get('ip'))
    if contar_cambios_usuario_mes(cliente['cedula']) >= limite_cambios:
        return jsonify({'success': False,
                        'message': f"Has alcanzado el límite de {limite_cambios} cambios permitidos este mes."})

    data = request.get_json()
    nueva_clave    = data.get("nueva_clave")
    confirmar_clave = data.get("confirmar_clave")

    if not nueva_clave or not confirmar_clave:
        return jsonify({'success': False, 'message': 'Por favor ingrese y confirme la nueva clave'})
    if nueva_clave != confirmar_clave:
        return jsonify({'success': False, 'message': 'Las contraseñas no coinciden'})
    if len(nueva_clave) < 8:
        return jsonify({'success': False, 'message': 'La contraseña debe tener al menos 8 caracteres'})

    device_info        = session.get('device_info', {})
    device_id          = device_info.get('device_id')
    interfaces_activas = device_info.get('interfaces_activas', [])

    if not device_id:
        # Re-validar automáticamente si la sesión no tiene device_info
        ip = session.get('ip')
        if ip:
            try:
                dispositivo_online, device_id = genieacs.obtener_estado_online_device(ip)
                if not dispositivo_online:
                    device_id = genieacs.obtener_device_id_por_ip(ip)
                if device_id:
                    interfaces_activas = genieacs.detectar_interfaces_wifi_activas(device_id)
                    session['device_info'] = {
                        'device_id': device_id,
                        'interfaces_activas': interfaces_activas,
                        'detected_at': datetime.now().isoformat()
                    }
            except Exception:
                pass
        if not device_id:
            return jsonify({'success': False, 'message': 'No se encontró el dispositivo. Verifica que esté conectado e intenta de nuevo.'})

    if accion == 'whatsapp':
        whatsapp = ycloud.normalizar_numero(cliente['celular'])
        if not whatsapp:
            return jsonify({'success': False, 'message': 'El número de teléfono no es válido para WhatsApp.'})
        if not ycloud.enviar_confirmacion_cambio(whatsapp, nueva_clave):
            return jsonify({'success': False, 'message': 'No se pudo enviar el WhatsApp.'})
        return jsonify({'success': True})

    if accion == 'cambiar':
        ok, mensaje = genieacs.cambiar_contraseña_wifi_interfaces_ya_detectadas(device_id, interfaces_activas, nueva_clave)
        if ok:
            wisphub.actualizar_parametros(cliente['cedula'], nueva_clave=nueva_clave, ip_especifica=session.get('ip'))
            registrar_cambio_usuario(cliente['cedula'], 'Password', 'Nueva')
            return jsonify({'success': True,
                            'message': f'{mensaje}. La ONU se reiniciará automáticamente para aplicar los cambios.'})
        return jsonify({'success': False, 'message': f'Error al cambiar la clave: {mensaje}'})

    return jsonify({'success': False, 'message': 'Acción no válida.'})


@app.route("/cambiar_nombre_red", methods=["GET", "POST"])
@cliente_requerido
def cambiar_nombre_red():
    limpiar_historial_antiguo()
    cliente = obtener_cliente_actual()

    if request.method == "GET":
        return render_template("users/user_cambiar_nombre_red.html", cliente=cliente)

    accion = request.args.get('accion')

    if accion == 'validar_dispositivo':
        ip = session.get('ip')
        if not ip:
            return jsonify({'success': False, 'message': 'No se encontró la IP del cliente'})
        try:
            dispositivo_online, device_id = genieacs.obtener_estado_online_device(ip)
            if not dispositivo_online:
                if not genieacs.verificar_dispositivo_online_alternativo(ip):
                    return jsonify({'success': False, 'message': 'El dispositivo está desconectado'})
                device_id = genieacs.obtener_device_id_por_ip(ip)
                if not device_id:
                    return jsonify({'success': False, 'message': 'No se encontró el dispositivo'})
            interfaces_activas, interfaces_info = genieacs.detectar_interfaces_y_frecuencias(device_id)
            if not interfaces_activas:
                return jsonify({'success': False, 'message': 'No se encontraron interfaces WiFi activas'})
            session['device_info_red'] = {
                'device_id': device_id,
                'interfaces_activas': interfaces_activas,
                'interfaces_info': interfaces_info,
                'detected_at': datetime.now().isoformat()
            }
            return jsonify({'success': True, 'device_id': device_id,
                            'interfaces_activas': interfaces_activas,
                            'interfaces_info': interfaces_info,
                            'interfaces_count': len(interfaces_activas)})
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error validando dispositivo: {e}'})

    if not request.is_json:
        return jsonify({'success': False, 'message': 'Solo se permite el flujo AJAX.'})

    limite_cambios = obtener_limite_cliente(session.get('ip'))
    if contar_cambios_usuario_mes(cliente['cedula']) >= limite_cambios:
        return jsonify({'success': False,
                        'message': f"Has alcanzado el límite de {limite_cambios} cambios permitidos este mes."})

    data = request.get_json()
    nuevo_nombre = data.get("nuevo_nombre")
    if not nuevo_nombre:
        return jsonify({'success': False, 'message': 'Por favor ingrese un nuevo nombre'})

    device_info     = session.get('device_info_red', {})
    device_id       = device_info.get('device_id')
    interfaces_info = device_info.get('interfaces_info', {})

    if not device_id:
        ip = session.get('ip')
        if ip:
            try:
                _, device_id = genieacs.obtener_estado_online_device(ip)
                if not device_id:
                    device_id = genieacs.obtener_device_id_por_ip(ip)
                if device_id:
                    interfaces_activas, interfaces_info = genieacs.detectar_interfaces_y_frecuencias(device_id)
                    session['device_info_red'] = {
                        'device_id': device_id,
                        'interfaces_activas': interfaces_activas,
                        'interfaces_info': interfaces_info,
                        'detected_at': datetime.now().isoformat()
                    }
            except Exception:
                pass
        if not device_id:
            return jsonify({'success': False, 'message': 'No se encontró el dispositivo. Verifica que esté conectado e intenta de nuevo.'})

    if accion == 'whatsapp':
        whatsapp = ycloud.normalizar_numero(cliente['celular'])
        if not whatsapp:
            return jsonify({'success': False, 'message': 'El número de teléfono no es válido para WhatsApp.'})
        if not ycloud.enviar_confirmacion_cambio(whatsapp, nuevo_nombre):
            return jsonify({'success': False, 'message': 'No se pudo enviar el WhatsApp.'})
        return jsonify({'success': True})

    if accion == 'cambiar':
        ok, mensaje = genieacs.cambiar_nombre_red_wifi_inteligente(device_id, interfaces_info, nuevo_nombre)
        if ok:
            wisphub.actualizar_parametros(cliente['cedula'], nuevo_ssid=nuevo_nombre, ip_especifica=session.get('ip'))
            registrar_cambio_usuario(cliente['cedula'], 'SSID', nuevo_nombre)
            return jsonify({'success': True,
                            'message': f'{mensaje}. La ONU se reiniciará automáticamente para aplicar los cambios.'})
        return jsonify({'success': False, 'message': f'Error al cambiar el nombre de la red: {mensaje}'})

    return jsonify({'success': False, 'message': 'Acción no válida.'})


@app.route('/reenviar_codigo', methods=['POST'])
def reenviar_codigo():
    telefono = session.get('telefono')
    codigo   = session.get('codigo_verificacion')
    if not telefono or not codigo:
        return jsonify({'success': False, 'message': 'No hay sesión activa.'})
    if not ycloud.enviar_otp(telefono, codigo):
        return jsonify({'success': False, 'message': 'No se pudo enviar el código.'})
    return jsonify({'success': True})


@app.route("/cerrar_sesion", methods=["POST"])
def cerrar_sesion():
    session.clear()
    flash("Sesión cerrada exitosamente", "success")
    return redirect(url_for("index"))


@app.route('/verificar_sesion', methods=['POST'])
def verificar_sesion():
    if not session.get("cliente_encontrado"):
        return jsonify({'error': 'Sesión expirada'}), 401
    return jsonify({'ok': True})


@app.route('/renovar_sesion', methods=['POST'])
def renovar_sesion():
    session.modified = True
    return '', 204


@app.route('/api/mis_dispositivos')
@cliente_requerido
def mis_dispositivos():
    ip = session.get('ip')
    if not ip:
        return jsonify({'success': False, 'count': 0, 'hosts': []})
    try:
        svc = genieacs._get_svc()
        device_id = genieacs.obtener_device_id_por_ip(ip)
        if not device_id:
            return jsonify({'success': False, 'count': 0, 'hosts': [], 'message': 'Dispositivo no encontrado'})
        device = svc.get_device_by_id(device_id)
        if not device:
            return jsonify({'success': False, 'count': 0, 'hosts': [], 'message': 'No se pudo leer el dispositivo'})
        all_hosts = svc.get_merged_hosts(device) or []
        active_hosts = [h for h in all_hosts if h.get('active', True)]
        return jsonify({
            'success': True,
            'online': svc.is_online(device),
            'count': len(active_hosts),
            'hosts': active_hosts
        })
    except Exception as e:
        return jsonify({'success': False, 'count': 0, 'hosts': [], 'error': str(e)})


# ===========================================================================
#  █████╗ ██████╗ ███╗   ███╗██╗███╗   ██╗
# ██╔══██╗██╔══██╗████╗ ████║██║████╗  ██║
# ███████║██║  ██║██╔████╔██║██║██╔██╗ ██║
# ██╔══██║██║  ██║██║╚██╔╝██║██║██║╚██╗██║
# ██║  ██║██████╔╝██║ ╚═╝ ██║██║██║ ╚████║
# ╚═╝  ╚═╝╚═════╝ ╚═╝     ╚═╝╚═╝╚═╝  ╚═══╝
# ===========================================================================


# ---------------------------------------------------------------------------
# Autenticación y decoradores admin
# ---------------------------------------------------------------------------

def requiere_api_key(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        api_key = os.getenv("SKYPASS_API_KEY")
        if not api_key:
            return jsonify({'success': False, 'message': 'API key no configurada en el servidor'}), 500
        if request.headers.get("X-Api-Key") != api_key:
            return jsonify({'success': False, 'message': 'API key inválida o ausente'}), 401
        return f(*args, **kwargs)
    return decorated


def admin_requerido(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_id' not in session:
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                      request.accept_mimetypes.best == 'application/json' or \
                      request.path.startswith('/api/') or \
                      'application/json' in (request.content_type or '') or \
                      request.method == 'POST'
            if is_ajax:
                return jsonify({'success': False, 'message': 'Sesión expirada. Inicia sesión de nuevo.'}), 401
            flash('Por favor inicie sesión como administrador', 'error')
            return redirect(url_for('admin.login'))
        return f(*args, **kwargs)
    return decorated_function


def verificar_admin(username, password):
    try:
        conn = get_db_connection()
        cur = conn.execute(
            "SELECT * FROM admin_users WHERE username = ? OR email = ?",
            (username, username)
        )
        admin = cur.fetchone()
        conn.close()
        if admin and check_password_hash(admin['password'], password):
            return admin
        return None
    except Exception as e:
        print(f"Error al verificar admin: {e}")
        return None


# ---------------------------------------------------------------------------
# Registro de cambios (admin)
# ---------------------------------------------------------------------------

def registrar_cambio(admin_id, cedula, tipo_cambio, valor_nuevo):
    try:
        conn = get_db_connection()
        conn.execute(
            "INSERT INTO change_history (admin_id, cedula, tipo_cambio, valor_nuevo, fecha) VALUES (?, ?, ?, ?, ?)",
            (admin_id, cedula, tipo_cambio, valor_nuevo, datetime.now().isoformat())
        )
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error al registrar cambio: {e}")
        return False


# ---------------------------------------------------------------------------
# Configuración y estadísticas (admin)
# ---------------------------------------------------------------------------

def obtener_estadisticas_uso():
    try:
        conn = get_db_connection()
        cambios_por_tipo = [dict(r) for r in conn.execute(
            "SELECT tipo_cambio, COUNT(*) as count FROM change_history GROUP BY tipo_cambio"
        ).fetchall()]
        cambios_por_mes = [dict(r) for r in conn.execute(
            "SELECT substr(fecha, 1, 7) as mes, COUNT(*) as count FROM change_history GROUP BY mes"
        ).fetchall()]
        limites_comunes = [dict(r) for r in conn.execute(
            "SELECT limite_personalizado, COUNT(*) as count FROM user_limits GROUP BY limite_personalizado"
        ).fetchall()]
        conn.close()
        return {"cambios_por_tipo": cambios_por_tipo,
                "cambios_por_mes": cambios_por_mes,
                "limites_comunes": limites_comunes}
    except Exception as e:
        print(f"Error al obtener estadísticas: {e}")
        return {}


def obtener_configuracion_global():
    try:
        conn = get_db_connection()
        data = [dict(r) for r in conn.execute("SELECT * FROM admin_settings").fetchall()]
        conn.close()
        return data
    except Exception as e:
        print(f"Error al obtener configuración global: {e}")
        return []


def actualizar_limite_cliente(ip, nombre, nuevo_limite, cedula=''):
    try:
        conn = get_db_connection()
        conn.execute("""
            INSERT INTO user_limits (ip, cedula, nombre, limite_personalizado)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(ip) DO UPDATE SET
                limite_personalizado=excluded.limite_personalizado,
                nombre=excluded.nombre,
                cedula=excluded.cedula;
        """, (ip, cedula, nombre, nuevo_limite))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error al actualizar límite: {e}")
        return False


# ---------------------------------------------------------------------------
# Rutas admin — autenticación
# ---------------------------------------------------------------------------

@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        if ip and ',' in ip:
            ip = ip.split(',')[0].strip()
            
        block_time = get_ip_block_time(ip)
        if block_time > 0:
            return jsonify({'success': False, 'message': f'IP bloqueada por intentos fallidos. Espera {block_time // 60}m {block_time % 60}s.'})

        if not request.is_json:
            return jsonify({'success': False, 'message': 'Solo se permite el flujo AJAX.'})
        data = request.get_json()
        result = verificar_admin(data.get('username'), data.get('password'))
        if result:
            clear_failed_attempts(ip)
            session.permanent = True
            session['admin_autenticado'] = True
            session['admin_id'] = result['id']
            session['admin_username'] = result['email']
            return jsonify({'success': True, 'redirect': url_for('admin.dashboard_admin')})
        
        record_failed_attempt(ip)
        return jsonify({'success': False, 'message': 'Credenciales inválidas'})
    return render_template('admin/admin_login.html')


@admin_bp.route('/logout', methods=['POST'])
def logout():
    session.pop('admin_autenticado', None)
    session.pop('admin_id', None)
    session.pop('admin_username', None)
    session.pop('cliente_encontrado', None)
    return jsonify({'success': True, 'redirect': url_for('admin.login')})


# ---------------------------------------------------------------------------
# Rutas admin — dashboard
# ---------------------------------------------------------------------------

@admin_bp.route('/dashboard', methods=['GET', 'POST'])
@admin_requerido
def dashboard_admin():
    return render_template(
        'admin/admin_dashboard.html',
        empresa=os.getenv("EMPRESA", "SkyPass"),
    )


# ---------------------------------------------------------------------------
# Rutas admin — historial
# ---------------------------------------------------------------------------

@admin_bp.route('/historial')
@admin_requerido
def historial():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    filtro_cedula = request.args.get('cedula', '').strip()
    filtro_tipo   = request.args.get('tipo_cambio', '').strip()

    where_clauses = []
    params = []
    if filtro_cedula:
        where_clauses.append("cedula = ?")
        params.append(filtro_cedula)
    if filtro_tipo:
        where_clauses.append("tipo_cambio = ?")
        params.append(filtro_tipo)

    where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    query = f"SELECT * FROM change_history{where_sql} ORDER BY fecha DESC LIMIT ? OFFSET ?"
    params_pag = params + [per_page, (page - 1) * per_page]

    count_query = f"SELECT COUNT(*) as total FROM change_history{where_sql}"

    conn = get_db_connection()
    historial_data = [dict(r) for r in conn.execute(query, params_pag).fetchall()]
    total_count    = conn.execute(count_query, params).fetchone()[0]
    tipos_cambio   = [r['tipo_cambio'] for r in
                      conn.execute("SELECT DISTINCT tipo_cambio FROM change_history ORDER BY tipo_cambio").fetchall()]
    conn.close()

    return render_template(
        'admin/admin_historial.html',
        historial=historial_data,
        current_page=page,
        total_pages=(total_count + per_page - 1) // per_page,
        filtro_cedula=filtro_cedula,
        filtro_tipo=filtro_tipo,
        tipos_cambio=tipos_cambio
    )


@admin_bp.route('/eliminar_cambio_historial', methods=['POST'])
@admin_requerido
def eliminar_cambio_historial():
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Solo se permite el flujo AJAX.'})
    id_cambio = request.get_json().get('id')
    if not id_cambio:
        return jsonify({'success': False, 'message': 'ID no proporcionado.'})
    try:
        conn = get_db_connection()
        conn.execute("DELETE FROM change_history WHERE id = ?", (id_cambio,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al eliminar el registro: {str(e)}'})


# ---------------------------------------------------------------------------
# Rutas admin — estadísticas y configuración
# ---------------------------------------------------------------------------

@admin_bp.route('/estadisticas')
@admin_requerido
def estadisticas():
    return render_template('admin/admin_estadisticas.html', estadisticas=obtener_estadisticas_uso())


@admin_bp.route('/configuracion', methods=['GET'])
@admin_requerido
def configuracion():
    try:
        conn = get_db_connection()
        data = conn.execute("SELECT valor FROM admin_settings WHERE clave = 'max_cambios_mes'").fetchone()
        conn.close()
        cambios_por_mes = int(data['valor']) if data else 2
    except Exception:
        cambios_por_mes = 2
    return render_template('admin/admin_configuracion.html', cambios_por_mes=cambios_por_mes)


@admin_bp.route('/cambiar_config', methods=['POST'])
@admin_requerido
def cambiar_config():
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Solo se permite el flujo AJAX.'})
    try:
        cambios_por_mes = int(request.get_json().get('cambios_por_mes'))
        if cambios_por_mes < 1:
            return jsonify({'success': False, 'message': 'El valor debe ser mayor a 0'})
        conn = get_db_connection()
        conn.execute("""
            INSERT INTO admin_settings (clave, valor) VALUES ('max_cambios_mes', ?)
            ON CONFLICT(clave) DO UPDATE SET valor=excluded.valor;
        """, (str(cambios_por_mes),))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Límite global actualizado correctamente.'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al actualizar la configuración global: {str(e)}'})


@admin_bp.route('/cambiar_password', methods=['POST'])
@admin_requerido
def cambiar_password():
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Solo se permite el flujo AJAX.'})
    data = request.get_json()
    nueva_password = data.get('nueva_password', '')
    if len(nueva_password) < 8:
        return jsonify({'success': False, 'message': 'La nueva contraseña debe tener al menos 8 caracteres'})
    if nueva_password != data.get('confirmar_password'):
        return jsonify({'success': False, 'message': 'Las contraseñas no coinciden'})
    try:
        conn = get_db_connection()
        admin_data = conn.execute("SELECT * FROM admin_users WHERE id = ?", (session['admin_id'],)).fetchone()
        if not admin_data or not check_password_hash(admin_data['password'], data.get('password_actual')):
            conn.close()
            return jsonify({'success': False, 'message': 'Contraseña actual incorrecta'})
        conn.execute("UPDATE admin_users SET password = ? WHERE id = ?",
                     (generate_password_hash(nueva_password), session['admin_id']))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Contraseña actualizada exitosamente'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al cambiar la contraseña: {str(e)}'})


# ---------------------------------------------------------------------------
# Rutas admin — límites personalizados
# ---------------------------------------------------------------------------

@admin_bp.route('/limites', methods=['GET', 'POST'])
@admin_requerido
def limites():
    if request.method == 'POST':
        ip = request.form.get('ip')
        nuevo_limite = int(request.form.get('nuevo_limite'))
        cliente = wisphub.buscar_por_ip(ip)
        if not cliente:
            flash('La IP ingresada no corresponde a ningún cliente.', 'error')
            return redirect(url_for('admin.limites'))
        conn = get_db_connection()
        existe = conn.execute("SELECT * FROM user_limits WHERE ip = ?", (ip,)).fetchone()
        conn.close()
        if actualizar_limite_cliente(ip, cliente.get('nombre', ''), nuevo_limite, cliente.get('cedula', '')):
            flash('Límite actualizado correctamente.' if existe else 'Límite creado correctamente.', 'personalizado')
        else:
            flash('Error al actualizar el límite personalizado.', 'error')
        return redirect(url_for('admin.limites'))

    try:
        conn = get_db_connection()
        limites_data    = [dict(r) for r in conn.execute("SELECT * FROM user_limits ORDER BY ip").fetchall()]
        data            = conn.execute("SELECT valor FROM admin_settings WHERE clave = 'max_cambios_mes'").fetchone()
        conn.close()
        cambios_por_mes = max(1, int(data['valor'])) if data and data['valor'] else 1
        return render_template('admin/admin_limites.html', limites=limites_data, cambios_por_mes=cambios_por_mes)
    except Exception as e:
        print(f'Error en /limites: {e}')
        flash('Error al cargar los límites', 'error')
        return render_template('admin/admin_limites.html', limites=[], cambios_por_mes=1)


@admin_bp.route('/buscar_cliente_limite', methods=['POST'])
@admin_requerido
def buscar_cliente_limite():
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Solo se permite el flujo AJAX.'})
    busqueda = request.get_json().get('busqueda', '').strip()
    if not busqueda:
        return jsonify({'success': False, 'message': 'Debe ingresar una IP o cédula'})
    try:
        partes = busqueda.split('.')
        if len(partes) == 4 and all(p.isdigit() for p in partes):
            cliente = wisphub.buscar_por_ip(busqueda)
        else:
            resultados = wisphub.buscar_por_cedula_parcial(busqueda)
            cedula_exacta = next((c for c in resultados if c.get('cedula') == busqueda), None)
            item = cedula_exacta or (resultados[0] if resultados else None)
            cliente, _ = wisphub.buscar_cliente(item['cedula'], timeout=7) if item else (None, None)

        if not cliente:
            return jsonify({'success': False, 'message': 'Cliente no encontrado'})

        ip_cliente = cliente.get('ip') or cliente.get('ip_address', '')
        if not ip_cliente:
            return jsonify({'success': False, 'message': 'El cliente no tiene IP asignada'})

        return jsonify({'success': True, 'cliente': {
            'ip': ip_cliente,
            'cedula': cliente.get('cedula', ''),
            'nombre': cliente.get('nombre', '')
        }})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al buscar cliente: {str(e)}'})


@admin_bp.route('/editar_limite', methods=['POST'])
@admin_requerido
def editar_limite():
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Solo se permite el flujo AJAX.'})
    data = request.get_json()
    ip = data.get('ip')
    cedula = data.get('cedula', '')
    nuevo_limite = data.get('nuevo_limite')
    if not ip or not nuevo_limite:
        return jsonify({'success': False, 'message': 'Datos incompletos para editar el límite'})
    try:
        nombre = ''
        cliente = wisphub.buscar_por_ip(ip)
        if cliente:
            nombre = cliente.get('nombre', '')
            cedula = cliente.get('cedula', cedula)
        conn = get_db_connection()
        conn.execute("""
            INSERT INTO user_limits (ip, cedula, nombre, limite_personalizado)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(ip) DO UPDATE SET
                limite_personalizado=excluded.limite_personalizado,
                nombre=excluded.nombre,
                cedula=excluded.cedula;
        """, (ip, cedula, nombre, int(nuevo_limite)))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Límite actualizado correctamente.',
                        'nuevo_limite': int(nuevo_limite)})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al actualizar el límite: {str(e)}'})


@admin_bp.route('/eliminar_limite', methods=['POST'])
@admin_requerido
def eliminar_limite():
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Solo se permite el flujo AJAX.'})
    ip = request.get_json().get('ip')
    if not ip:
        return jsonify({'success': False, 'message': 'IP no proporcionada'})
    try:
        conn = get_db_connection()
        conn.execute("DELETE FROM user_limits WHERE ip = ?", (ip,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Límite personalizado eliminado.'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al eliminar el límite: {str(e)}'})


# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Rutas admin — búsqueda de clientes (autocomplete)
# ---------------------------------------------------------------------------

@admin_bp.route('/api/buscar_clientes')
def api_buscar_clientes():
    q = request.args.get('q', '').strip()
    tipo_busqueda = request.args.get('tipo', 'nombre').strip()

    if not q:
        return jsonify({'success': True, 'clientes': []})
    try:
        if tipo_busqueda == 'ip':
            cliente = wisphub.buscar_por_ip(q)
            clientes = [{
                'ip':       cliente.get('ip') or cliente.get('ip_address') or '',
                'nombre':   cliente.get('nombre', ''),
                'cedula':   cliente.get('cedula', ''),
                'telefono': cliente.get('telefono', '') or cliente.get('celular', ''),
                'estado':   cliente.get('estado', ''),
            }] if cliente else []
        elif tipo_busqueda == 'cedula':
            clientes = wisphub.buscar_por_cedula_parcial(q)
        else:
            clientes = wisphub.buscar_por_nombre_parcial(q)
        return jsonify({'success': True, 'clientes': clientes})
    except Exception as e:
        print(f'Error en api_buscar_clientes: {e}')
        return jsonify({'success': False, 'clientes': []})


@admin_bp.route('/api/device_id_por_ip')
@admin_requerido
def api_device_id_por_ip():
    ip = request.args.get('ip', '').strip()
    if not ip:
        return jsonify({'found': False})
    try:
        device_id = genieacs.obtener_device_id_por_ip(ip)
        if device_id:
            return jsonify({'found': True, 'device_id': device_id})
        return jsonify({'found': False})
    except Exception:
        return jsonify({'found': False})


@admin_bp.route('/api/acs_status')
@admin_requerido
def api_acs_status():
    """Devuelve estado ACS (online/offline/no_encontrado) para una IP."""
    ip = request.args.get('ip', '').strip()
    if not ip:
        return jsonify({'found': False})
    try:
        svc    = genieacs._get_svc()
        device = svc.get_device_by_ip(ip)
        if not device:
            return jsonify({'found': False})
        return jsonify({
            'found':     True,
            'online':    svc.is_online(device),
            'device_id': device.get('_id', ''),
        })
    except Exception:
        return jsonify({'found': False})


@admin_bp.route('/api/ping')
@admin_requerido
def api_ping():
    """Ejecuta ping desde GenieACS y devuelve el avg en ms."""
    import re
    import requests as req
    ip = request.args.get('ip', '').strip()
    if not ip:
        return jsonify({'success': False, 'error': 'IP requerida'})
    try:
        genieacs_url = os.getenv('GENIEACS_API_URL', 'http://localhost:7557')
        resp = req.get(f'{genieacs_url}/ping/{ip}', timeout=15)
        text = resp.text
        match = re.search(r'rtt min/avg/max/mdev = [\d.]+/([\d.]+)/[\d.]+/[\d.]+ ms', text)
        loss_match = re.search(r'(\d+)% packet loss', text)
        loss = int(loss_match.group(1)) if loss_match else 100
        if match:
            return jsonify({'success': True, 'avg': round(float(match.group(1)), 1), 'loss': loss})
        return jsonify({'success': True, 'avg': None, 'loss': loss})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@admin_bp.route('/clientes')
@admin_requerido
def clientes():
    return render_template('admin/clientes.html')


@admin_bp.route('/api/wifi_password', methods=['POST'])
@admin_requerido
def api_wifi_password():
    """Cambia la contraseña WiFi de un dispositivo por device_id."""
    data          = request.get_json(silent=True) or {}
    device_id     = (data.get('device_id') or '').strip()
    nueva_password = (data.get('nueva_password') or '').strip()

    if not device_id:
        return jsonify({'success': False, 'message': 'device_id requerido'}), 400
    if len(nueva_password) < 8:
        return jsonify({'success': False, 'message': 'La contraseña debe tener al menos 8 caracteres'}), 400

    interfaces = genieacs.detectar_interfaces_wifi_activas(device_id)
    if not interfaces:
        return jsonify({'success': False, 'message': 'No se encontraron interfaces WiFi en el dispositivo'})

    ok, msg = genieacs.cambiar_contraseña_wifi_interfaces_ya_detectadas(device_id, interfaces, nueva_password)
    return jsonify({'success': ok, 'message': msg})


@admin_bp.route('/api/clientes')
@admin_requerido
def api_clientes():
    page          = max(1, int(request.args.get('page', 1)))
    page_size     = min(int(request.args.get('page_size', 20)), 50)
    q             = request.args.get('q', '').strip()
    tipo          = request.args.get('tipo', '').strip()
    force         = request.args.get('force_refresh', '0') == '1'
    estado_filter = request.args.get('estado_filter', '').strip().lower()

    # Búsqueda server-side en WispHub cuando hay tipo + q
    if q and tipo in ('nombre', 'cedula', 'ip'):
        if tipo == 'nombre':
            todos = wisphub.buscar_por_nombre_parcial(q)
        elif tipo == 'cedula':
            todos = wisphub.buscar_por_cedula_parcial(q)
        else:
            todos = wisphub.buscar_por_ip_parcial(q)
    else:
        todos = wisphub.listar_todos(force_refresh=force)
        if q:
            q_lower = q.lower()
            tokens  = q_lower.split()
            def matches(c):
                haystack = ' '.join([
                    (c.get('nombre')   or '').lower(),
                    (c.get('cedula')   or '').lower(),
                    (c.get('ip')       or '').lower(),
                    (c.get('telefono') or '').lower(),
                    (c.get('estado')   or '').lower(),
                ])
                return all(t in haystack for t in tokens)
            todos = [c for c in todos if matches(c)]

    if estado_filter:
        todos = [c for c in todos if (c.get('estado') or '').lower() == estado_filter]

    total       = len(todos)
    total_pages = max(1, -(-total // page_size))
    page        = min(page, total_pages)
    start       = (page - 1) * page_size
    pagina      = todos[start: start + page_size]

    return jsonify({
        'success':     True,
        'clientes':    pagina,
        'total':       total,
        'page':        page,
        'page_size':   page_size,
        'total_pages': total_pages,
    })


@admin_bp.route('/api/clientes/exportar')
@admin_requerido
def exportar_clientes_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q     = request.args.get('q', '').strip().lower()
    force = request.args.get('force_refresh', '0') == '1'

    todos = wisphub.listar_todos(force_refresh=force)

    if q:
        tokens = q.split()
        def matches(c):
            haystack = ' '.join([
                (c.get('nombre')   or '').lower(),
                (c.get('cedula')   or '').lower(),
                (c.get('ip')       or '').lower(),
                (c.get('telefono') or '').lower(),
                (c.get('estado')   or '').lower(),
            ])
            return all(t in haystack for t in tokens)
        todos = [c for c in todos if matches(c)]

    # ── Mapa ACS: IP → {'found': bool, 'online': bool} ──
    acs_map = {}
    try:
        svc = genieacs._get_svc()
        # Traer todos los dispositivos con proyección mínima
        dispositivos = svc.get_all_devices(projection=(
            '_id,_lastInform,'
            'VirtualParameters.IPTR069._value,'
            'InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANIPConnection.1.ExternalIPAddress._value,'
            'InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANPPPConnection.1.ExternalIPAddress._value,'
            'InternetGatewayDevice.ManagementServer.ConnectionRequestURL._value'
        ))

        def _ip_de_device(d):
            for path in [
                'VirtualParameters.IPTR069._value',
                'InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANIPConnection.1.ExternalIPAddress._value',
                'InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANPPPConnection.1.ExternalIPAddress._value',
            ]:
                parts = path.split('.')
                node = d
                try:
                    for p in parts:
                        node = node[p]
                    if node:
                        return str(node)
                except (KeyError, TypeError):
                    pass
            # Fallback: extraer IP de ConnectionRequestURL
            import re
            url = ''
            try:
                url = d['InternetGatewayDevice']['ManagementServer']['ConnectionRequestURL']['_value'] or ''
            except (KeyError, TypeError):
                pass
            m = re.search(r'http://([\d.]+):', url)
            return m.group(1) if m else None

        for d in dispositivos:
            ip = _ip_de_device(d)
            if ip:
                acs_map[ip] = {
                    'found':  True,
                    'online': svc.is_online(d),
                }
    except Exception:
        pass  # Si ACS no responde, la columna queda como 'Sin datos'

    # ── Libro Excel ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Clientes'

    azul       = '2563EB'
    azul_claro = 'EFF6FF'
    gris_borde = 'E5E7EB'

    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor=azul)
    header_align = Alignment(horizontal='center', vertical='center')

    row_font_bold = Font(name='Calibri', bold=True, size=10)
    row_font      = Font(name='Calibri', size=10)
    row_mono      = Font(name='Courier New', size=9.5)
    row_align     = Alignment(vertical='center', horizontal='left')
    center_align  = Alignment(vertical='center', horizontal='center')

    thin  = Side(style='thin', color=gris_borde)
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_online  = PatternFill('solid', fgColor='DCFCE7')
    fill_offline = PatternFill('solid', fgColor='FEE2E2')
    fill_noacs   = PatternFill('solid', fgColor='F3F4F6')
    fill_alt     = PatternFill('solid', fgColor='F9FAFB')

    font_online  = Font(name='Calibri', bold=True, size=10, color='15803D')
    font_offline = Font(name='Calibri', bold=True, size=10, color='B91C1C')
    font_noacs   = Font(name='Calibri', size=10,            color='6B7280')

    # Encabezados
    columnas = ['#', 'Nombre', 'Cédula', 'Teléfono', 'IP', 'Estado ACS', 'Fecha exportación']
    anchos   = [5,   35,       18,       16,          18,   16,            22]

    ws.row_dimensions[1].height = 28
    for col_idx, (titulo, ancho) in enumerate(zip(columnas, anchos), start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = borde
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    # Filas de datos
    fecha_export = datetime.now().strftime('%d/%m/%Y %H:%M')
    for fila_idx, cliente in enumerate(todos, start=2):
        ws.row_dimensions[fila_idx].height = 20
        use_alt = fila_idx % 2 == 0

        ip  = cliente.get('ip') or ''
        acs = acs_map.get(ip)

        if not ip or acs is None:
            acs_texto = 'Sin ACS'
            acs_font  = font_noacs
            acs_fill  = fill_noacs
        elif acs['online']:
            acs_texto = 'Online'
            acs_font  = font_online
            acs_fill  = fill_online
        else:
            acs_texto = 'Offline'
            acs_font  = font_offline
            acs_fill  = fill_offline

        filas_datos = [
            (fila_idx - 1,                       row_font,      row_align,  None),
            (cliente.get('nombre')   or '—',     row_font_bold, row_align,  None),
            (cliente.get('cedula')   or '—',     row_mono,      row_align,  None),
            (cliente.get('telefono') or '—',     row_font,      row_align,  None),
            (ip or '—',                          row_mono,      row_align,  None),
            (acs_texto,                          acs_font,      center_align, acs_fill),
            (fecha_export,                       row_font,      center_align, None),
        ]

        for col_idx, (valor, fuente, align, celda_fill) in enumerate(filas_datos, start=1):
            cell = ws.cell(row=fila_idx, column=col_idx, value=valor)
            cell.font      = fuente
            cell.alignment = align
            cell.border    = borde
            # La celda ACS tiene su propio color; el resto usa la alternancia
            if celda_fill:
                cell.fill = celda_fill
            elif use_alt:
                cell.fill = fill_alt

    # Fila totales
    total_row = len(todos) + 2
    ws.row_dimensions[total_row].height = 20
    online_count  = sum(1 for c in todos if acs_map.get(c.get('ip') or '', {}).get('online'))
    offline_count = sum(1 for c in todos if c.get('ip') and acs_map.get(c.get('ip'), {}).get('found') and not acs_map[c['ip']]['online'])
    resumen = f'Total: {len(todos)} clientes  |  ACS Online: {online_count}  |  ACS Offline: {offline_count}'
    total_cell = ws.cell(row=total_row, column=1, value=resumen)
    total_cell.font      = Font(name='Calibri', bold=True, size=10, color=azul)
    total_cell.fill      = PatternFill('solid', fgColor=azul_claro)
    total_cell.border    = borde
    total_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=len(columnas))

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre_archivo = f"clientes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre_archivo,
    )


@admin_bp.route('/buscar-cliente')
@admin_requerido
def buscar_cliente():
    return render_template('admin/buscar_cliente.html')


# ---------------------------------------------------------------------------
# Rutas admin — dispositivos conectados (GenieACS)
# ---------------------------------------------------------------------------

@admin_bp.route('/dispositivos_conectados')
@admin_requerido
def admin_dispositivos_conectados():
    return render_template('admin/dispositivos_conectados.html')


@admin_bp.route('/api/dispositivos_conectados')
@admin_requerido
def api_dispositivos_conectados():
    import re
    try:
        svc = genieacs._get_svc()
        projection = ",".join([
            "_id", "_lastInform", "_deviceId",
            "InternetGatewayDevice.DeviceInfo.ModelName",
            "InternetGatewayDevice.DeviceInfo.Manufacturer",
            "InternetGatewayDevice.DeviceInfo.SerialNumber",
            "InternetGatewayDevice.DeviceInfo.UpTime",
            "InternetGatewayDevice.ManagementServer.ConnectionRequestURL",
            "VirtualParameters.IPTR069",
            "VirtualParameters.PonMac",
            "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANIPConnection.1.ExternalIPAddress",
            "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANIPConnection.1.MACAddress",
            "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANPPPConnection.1.ExternalIPAddress",
            "InternetGatewayDevice.WANDevice.1.WANCommonInterfaceConfig.MACAddress",
            "InternetGatewayDevice.LANDevice.1.LANEthernetInterfaceConfig.1.MACAddress",
            "InternetGatewayDevice.LANDevice.1.WLANConfiguration.1.SSID",
            "InternetGatewayDevice.LANDevice.1.WLANConfiguration.5.SSID",
            "VirtualParameters.activedevices",
            "VirtualParameters.RXPower",
            "VirtualParameters.gettemp",
        ])
        devices = svc.get_all_devices(projection=projection)
        lista = []
        for device in devices:
            ip_actual = (
                svc._get_param(device, "VirtualParameters.IPTR069") or
                svc._get_param(device, "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANIPConnection.1.ExternalIPAddress") or
                svc._get_param(device, "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANPPPConnection.1.ExternalIPAddress") or ''
            )
            if not ip_actual:
                url_conn = svc._get_param(device, "InternetGatewayDevice.ManagementServer.ConnectionRequestURL") or ''
                m = re.search(r'http://([\d\.]+):', url_conn)
                if m:
                    ip_actual = m.group(1)
            dev_id  = device.get('_id', '')
            meta    = device.get('_deviceId') or {}
            if not isinstance(meta, dict):
                meta = {}
            serial  = meta.get('_SerialNumber') or (dev_id.split('-', 2)[2] if dev_id.count('-') >= 2 else dev_id)
            rx      = svc.get_rx_power(device)
            temp    = svc.get_temperature(device)
            ut      = svc.get_uptime_seconds(device)
            lista.append({
                'device_id':      dev_id,
                'serial':         serial or dev_id,
                'ip':             ip_actual,
                'mac':            svc.get_mac(device) or '',
                'ssid':           svc.get_ssid_primary(device) or '',
                'ultimo_informe': device.get('_lastInform'),
                'online':         svc.is_online(device),
                'tiempo':         svc.tiempo_desde_ultimo_inform(device),
                'modelo':         svc._get_param(device, "InternetGatewayDevice.DeviceInfo.ModelName") or meta.get('_ProductClass') or '',
                'fabricante':     svc._get_param(device, "InternetGatewayDevice.DeviceInfo.Manufacturer") or '',
                'uptime':         svc.format_uptime(ut),
                'rx':             rx,
                'rx_calidad':     svc.rx_signal_quality(rx),
                'temp':           temp,
                'hosts':          svc._get_param(device, "VirtualParameters.activedevices") or '0',
                'detras_nat':     svc.is_behind_nat(device),
            })
        return jsonify({'success': True, 'dispositivos': lista})
    except Exception as e:
        return jsonify({'success': False, 'dispositivos': [], 'error': str(e)})


@admin_bp.route('/api/dashboard_stats')
@admin_requerido
def api_dashboard_stats():
    import re
    result = {}
    try:
        conn = get_db_connection()
        ahora = datetime.now()
        inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        row = conn.execute(
            "SELECT COUNT(*) as total FROM change_history WHERE fecha >= ?",
            (inicio_mes.isoformat(),)
        ).fetchone()
        result['cambios_mes'] = row['total'] if row else 0
        recent_ch = conn.execute(
            "SELECT cedula, tipo_cambio, valor_nuevo, fecha FROM change_history ORDER BY fecha DESC LIMIT 8"
        ).fetchall()
        result['cambios_recientes'] = [dict(r) for r in recent_ch]
        conn.close()
    except Exception:
        result['cambios_mes'] = 0
        result['cambios_recientes'] = []
    try:
        svc = genieacs._get_svc()
        proj = ",".join([
            "_id", "_lastInform", "_deviceId",
            "InternetGatewayDevice.DeviceInfo.ModelName",
            "InternetGatewayDevice.DeviceInfo.UpTime",
            "InternetGatewayDevice.ManagementServer.ConnectionRequestURL",
            "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANIPConnection.1.ExternalIPAddress",
            "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANPPPConnection.1.ExternalIPAddress",
            "VirtualParameters.RXPower",
            "VirtualParameters.gettemp",
        ])
        devices = svc.get_all_devices(projection=proj)
        online_c = offline_c = 0
        uptimes = []
        rx_ok = rx_warn = rx_crit = 0
        recent_devs = []
        for d in devices:
            is_on = svc.is_online(d)
            if is_on:
                online_c += 1
            else:
                offline_c += 1
            ut = svc.get_uptime_seconds(d)
            if ut:
                uptimes.append(ut)
            rx = svc.get_rx_power(d)
            if rx is not None:
                q = svc.rx_signal_quality(rx)
                if q == "ok":      rx_ok   += 1
                elif q == "warning": rx_warn += 1
                else:              rx_crit += 1
            ip = (
                svc._get_param(d, "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANIPConnection.1.ExternalIPAddress") or
                svc._get_param(d, "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANPPPConnection.1.ExternalIPAddress") or ''
            )
            if not ip:
                url_c = svc._get_param(d, "InternetGatewayDevice.ManagementServer.ConnectionRequestURL") or ''
                m = re.search(r'http://([\d\.]+):', url_c)
                if m:
                    ip = m.group(1)
            recent_devs.append({
                'device_id': d.get('_id'),
                'ip': ip,
                'modelo': svc._get_param(d, "InternetGatewayDevice.DeviceInfo.ModelName") or '',
                'online': is_on,
                'tiempo': svc.tiempo_desde_ultimo_inform(d),
                'rx': rx,
                'ultimo_informe': d.get('_lastInform') or '',
            })
        recent_devs.sort(key=lambda x: x['ultimo_informe'], reverse=True)
        result.update({
            'devices_online':    online_c,
            'devices_offline':   offline_c,
            'uptime_promedio_h': round(sum(uptimes) / len(uptimes) / 3600, 1) if uptimes else None,
            'rx_ok':   rx_ok,
            'rx_warn': rx_warn,
            'rx_crit': rx_crit,
            'rx_total': rx_ok + rx_warn + rx_crit,
            'recent_devices': recent_devs[:6],
            'success': True,
        })
    except RuntimeError as e:
        result.update({
            'devices_online': 0, 'devices_offline': 0,
            'uptime_promedio_h': None,
            'rx_ok': 0, 'rx_warn': 0, 'rx_crit': 0, 'rx_total': 0,
            'recent_devices': [],
            'success': False, 'error': str(e),
        })
    return jsonify(result)


@admin_bp.route('/impersonate_client/<path:ip_wan>')
@admin_requerido
def impersonate_client(ip_wan):
    cliente = wisphub.buscar_por_ip(ip_wan)
    if not cliente:
        flash('No se pudo encontrar al cliente con la IP especificada en Wisphub.', 'error')
        return redirect(request.referrer or url_for('admin.admin_dashboard'))
    
    session.permanent = True
    session['telefono'] = cliente.get('telefono', '') or cliente.get('celular', '')
    session['nombre'] = cliente.get('nombre', '')
    session['cedula'] = cliente.get('cedula', '')
    session['ip'] = ip_wan
    session['cliente_encontrado'] = True
    
    flash(f'Sesión iniciada como {cliente.get("nombre", "Cliente")} (Impersonación)', 'success')
    return redirect(url_for('dashboard'))

@admin_bp.route('/dispositivo/<path:device_id>')
@admin_requerido
def admin_device_detail(device_id):
    return render_template('admin/admin_device_detail.html', device_id=device_id)


@admin_bp.route('/api/dispositivo/<path:device_id>')
@admin_requerido
def api_device_detail(device_id):
    try:
        svc    = genieacs._get_svc()
        device = svc.get_device_by_id(device_id)
        if not device:
            return jsonify({'success': False, 'error': 'Dispositivo no encontrado'})
        dev_id = device.get('_id', '')
        meta   = device.get('_deviceId') or {}
        if not isinstance(meta, dict):
            meta = {}
        serial        = meta.get('_SerialNumber') or (dev_id.split('-', 2)[2] if dev_id.count('-') >= 2 else dev_id)
        oui           = meta.get('_OUI')           or dev_id.split('-')[0]
        product_class = meta.get('_ProductClass')  or (dev_id.split('-', 2)[1] if dev_id.count('-') >= 2 else '')
        ip_wan = (
            svc._get_param(device, "VirtualParameters.IPTR069") or
            svc._get_param(device, "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANIPConnection.1.ExternalIPAddress") or
            svc._get_param(device, "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANPPPConnection.1.ExternalIPAddress") or ''
        )
        wan_status = (
            svc._get_param(device, "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANIPConnection.1.ConnectionStatus") or
            svc._get_param(device, "InternetGatewayDevice.WANDevice.1.WANConnectionDevice.1.WANPPPConnection.1.ConnectionStatus") or ''
        )
        rx    = svc.get_rx_power(device)
        temp  = svc.get_temperature(device)
        ut    = svc.get_uptime_seconds(device)
        
        cliente_wisphub = wisphub.buscar_por_ip(ip_wan) if ip_wan else None
        cambios_recientes = []
        if cliente_wisphub and cliente_wisphub.get('cedula'):
            try:
                conn = get_db_connection()
                rows = conn.execute(
                    "SELECT tipo_cambio, valor_nuevo, fecha FROM change_history WHERE cedula = ? ORDER BY fecha DESC LIMIT 5",
                    (cliente_wisphub['cedula'],)
                ).fetchall()
                cambios_recientes = [dict(r) for r in rows]
                conn.close()
            except Exception as e:
                print(f"Error fetching change history: {e}")
        
        ifaces = svc.get_wifi_interfaces(device)
        ifaces_out = []
        for iface in ifaces:
            beacon = iface.get("security", "") or ""
            sec_map = {"11i": "WPA2", "WPAand11i": "WPA/WPA2", "WPA": "WPA", "None": "Abierta"}
            security = sec_map.get(beacon, beacon) or "WPA/WPA2"
            ifaces_out.append({
                "index":     iface["index"],
                "ssid":      iface["ssid"],
                "channel":   iface["channel"],
                "frequency": iface["frequency"],
                "security":  security,
            })
        return jsonify({
            'success': True,
            'device': {
                'id':             dev_id,
                'serial':         serial,
                'oui':            oui,
                'product_class':  product_class,
                'fabricante':     svc._get_param(device, "InternetGatewayDevice.DeviceInfo.Manufacturer") or meta.get('_Manufacturer', ''),
                'modelo':         svc._get_param(device, "InternetGatewayDevice.DeviceInfo.ModelName") or product_class,
                'firmware':       svc._get_param(device, "InternetGatewayDevice.DeviceInfo.SoftwareVersion") or '',
                'hardware':       svc._get_param(device, "InternetGatewayDevice.DeviceInfo.HardwareVersion") or '',
                'mac_wan':        svc.get_mac(device) or '',
                'ip_wan':         ip_wan,
                'wan_status':     wan_status,
                'conn_request_url': svc._get_param(device, "InternetGatewayDevice.ManagementServer.ConnectionRequestURL") or '',
                'uptime':         svc.format_uptime(ut),
                'ultimo_informe': device.get('_lastInform'),
                'online':         svc.is_online(device),
                'tiempo':         svc.tiempo_desde_ultimo_inform(device),
                'detras_nat':     svc.is_behind_nat(device),
                'rx_power':       rx,
                'cliente': {
                    'nombre': cliente_wisphub.get('nombre') if cliente_wisphub else None,
                    'cedula': cliente_wisphub.get('cedula') if cliente_wisphub else None,
                    'telefono': cliente_wisphub.get('telefono') or cliente_wisphub.get('celular') if cliente_wisphub else None,
                    'plan': cliente_wisphub.get('plan_internet', {}).get('nombre') if cliente_wisphub and cliente_wisphub.get('plan_internet') else None,
                } if cliente_wisphub else None,
                'cambios_recientes': cambios_recientes,
                'rx_calidad':     svc.rx_signal_quality(rx),
                'temperatura':    temp,
                'hosts_count':    svc._get_param(device, "VirtualParameters.activedevices") or '0',
                'wifi_interfaces': ifaces_out,
                'lan_hosts':      svc.get_merged_hosts(device),
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@admin_bp.route('/eliminar_device', methods=['POST'])
@admin_requerido
def eliminar_device():
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Solo se permite el flujo AJAX.'})
    data = request.get_json()
    device_id = data.get('device_id')
    ip = data.get('ip')
    if not device_id and not ip:
        return jsonify({'success': False, 'message': 'Se requiere device_id o IP del dispositivo.'})
    if not device_id and ip:
        device_id = genieacs.obtener_device_id_por_ip(ip)
        if not device_id:
            return jsonify({'success': False, 'message': f'No se encontró dispositivo con IP: {ip}'})
    if genieacs.eliminar_device_genieacs(device_id):
        return jsonify({'success': True, 'message': 'Dispositivo eliminado exitosamente.'})
    return jsonify({'success': False, 'message': 'Error al eliminar el dispositivo.'})


@admin_bp.route('/summon_device', methods=['POST'])
@admin_requerido
def summon_device():
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Solo se permite el flujo AJAX.'})
    try:
        device_id = request.json.get('device_id')
        if not device_id:
            return jsonify({'success': False, 'message': 'Falta device_id'})
        
        svc = genieacs._get_svc()
        # Una tarea refreshObject con connection_request=True fuerza la reconexión (summon)
        r = svc._post_task(device_id, {"name": "refreshObject", "objectName": ""}, connection_request=True)
        return jsonify({'success': r.get('success', False)})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ---------------------------------------------------------------------------
# Rutas admin — WhatsApp / YCloud
# ---------------------------------------------------------------------------

@admin_bp.route('/conectar-whatsapp')
@admin_requerido
def conectar_whatsapp():
    try:
        estado = 'conectado' if ycloud.verificar_estado() == 'CONNECTED' else 'desconectado'
    except Exception:
        estado = 'error'
    return render_template('admin/conectar_whatsapp.html', estado=estado, qr_url=None)


@admin_bp.route('/estado-bot', methods=['GET'])
@admin_requerido
def estado_bot():
    try:
        estado_ycloud = ycloud.verificar_estado()
        if estado_ycloud == 'CONNECTED':
            return jsonify({'conectado': True, 'estado': 'conectado'})
        return jsonify({'conectado': False, 'estado': 'desconectado', 'detalle': estado_ycloud})
    except Exception as e:
        return jsonify({'conectado': False, 'estado': 'error', 'error': str(e)})


# ---------------------------------------------------------------------------
# Rutas admin — sesión
# ---------------------------------------------------------------------------

@admin_bp.route('/verificar-sesion', methods=['POST'])
def verificar_sesion_admin():
    if 'admin_id' not in session:
        return jsonify({'error': 'Sesión expirada'}), 401
    return jsonify({'ok': True})


@admin_bp.route('/renovar-sesion', methods=['POST'])
@admin_requerido
def renovar_sesion_admin():
    session.modified = True
    return '', 204


# ===========================================================================
# Handlers globales
# ===========================================================================

@app.errorhandler(404)
def page_not_found(e):
    return render_template("404.html"), 404


@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    
    # Cabeceras de seguridad estrictas
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    
    return response


@admin_bp.route('/api/limpiar-duplicados', methods=['POST'])
@admin_requerido
def limpiar_duplicados_acs():
    try:
        resultado = genieacs._get_svc().limpiar_duplicados_ip()
        return jsonify({'success': True, **resultado})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


app.register_blueprint(admin_bp)


# ---------------------------------------------------------------------------
# API pública — autenticada con X-Api-Key
# ---------------------------------------------------------------------------

@app.route('/api/cambiar_wifi', methods=['POST'])
@requiere_api_key
def api_cambiar_wifi():
    """
    Body JSON:
      { "ip": "1.2.3.4", "nueva_password": "MiClave123" }
    Header:
      X-Api-Key: <SKYPASS_API_KEY>
    """
    data = request.get_json(silent=True) or {}
    ip            = (data.get('ip') or '').strip()
    nueva_password = (data.get('nueva_password') or '').strip()

    if not ip:
        return jsonify({'success': False, 'message': 'El campo "ip" es obligatorio'}), 400
    if not nueva_password:
        return jsonify({'success': False, 'message': 'El campo "nueva_password" es obligatorio'}), 400
    if len(nueva_password) < 8:
        return jsonify({'success': False, 'message': 'La contraseña debe tener al menos 8 caracteres'}), 400

    # 1. Buscar cliente en WispHub para obtener cédula y validar límite
    cliente = wisphub.buscar_por_ip(ip)
    cedula  = cliente.get('cedula') if cliente else None

    if cedula:
        limite_cambios   = obtener_limite_cliente(ip)
        cambios_realizados = contar_cambios_usuario_mes(cedula)
        if cambios_realizados >= limite_cambios:
            return jsonify({
                'success': False,
                'code':    'limite_alcanzado',
                'message': f'Límite de cambios alcanzado. Llevas {cambios_realizados} de {limite_cambios} cambios permitidos este mes.',
                'cambios_realizados': cambios_realizados,
                'limite':             limite_cambios,
            })

    # 2. Buscar el dispositivo en el servidor ACS
    svc = genieacs._get_svc()
    device = svc.get_device_by_ip(ip)
    if not device:
        return jsonify({'success': False, 'message': f'No se encontró ningún dispositivo con IP {ip}'}), 404

    device_id = device.get('_id')

    # 3. Detectar interfaces WiFi activas
    interfaces_activas = genieacs.detectar_interfaces_wifi_activas(device_id)
    if not interfaces_activas:
        return jsonify({'success': False, 'message': 'No se encontraron interfaces WiFi activas en el dispositivo'}), 422

    # 4. Cambiar la contraseña en cada interfaz y reiniciar
    ok, mensaje = genieacs.cambiar_contraseña_wifi_interfaces_ya_detectadas(
        device_id, interfaces_activas, nueva_password
    )
    if not ok:
        return jsonify({'success': False, 'message': mensaje}), 500

    if cedula:
        registrar_cambio_usuario(cedula, 'Password', 'Nueva')

    return jsonify({
        'success': True,
        'device_id': device_id,
        'interfaces': interfaces_activas,
        'message': mensaje,
    })

@app.route('/api/verificar_dispositivo', methods=['GET'])
@requiere_api_key
def api_verificar_dispositivo():
    """
    Verifica si una IP existe en GenieACS y si está online.

    Header: X-Api-Key: <SKYPASS_API_KEY>
    Query:  GET /api/verificar_dispositivo?ip=1.2.3.4

    Respuesta:
      { "success": true, "ip": "...", "encontrado": true/false, "online": true/false, "device_id": "..." }
    """
    ip = request.args.get('ip', '').strip()

    if not ip:
        return jsonify({'success': False, 'message': 'El parámetro "ip" es obligatorio'}), 400

    # Validación básica de formato IP
    partes = ip.split('.')
    if len(partes) != 4 or not all(p.isdigit() and 0 <= int(p) <= 255 for p in partes):
        return jsonify({'success': False, 'message': 'Formato de IP inválido'}), 400

    try:
        svc    = genieacs._get_svc()
        device = svc.get_device_by_ip(ip)

        if not device:
            return jsonify({'success': True, 'ip': ip, 'encontrado': False, 'online': False, 'device_id': None})

        online    = svc.is_online(device)
        device_id = device.get('_id')

        return jsonify({'success': True, 'ip': ip, 'encontrado': True, 'online': online, 'device_id': device_id})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error consultando GenieACS: {str(e)}'}), 500


# ---------------------------------------------------------------------------
# Scheduler — limpieza automática de duplicados ACS cada 30 min
# ---------------------------------------------------------------------------

def _job_limpiar_duplicados():
    INTERVALO = 30 * 60
    while True:
        time.sleep(INTERVALO)
        try:
            resultado = genieacs._get_svc().limpiar_duplicados_ip()
            eliminados = resultado.get('eliminados', [])
            if eliminados:
                print(f"[ACS] Limpieza duplicados: {len(eliminados)} eliminado(s) — {[e['ip'] for e in eliminados]}")
            else:
                print(f"[ACS] Limpieza duplicados: sin duplicados encontrados.")
        except Exception as e:
            print(f"[ACS] Error en limpieza de duplicados: {e}")

_scheduler = threading.Thread(target=_job_limpiar_duplicados, daemon=True, name="acs-limpiar-dup")
_scheduler.start()


if __name__ == "__main__":
    app.run(debug=False)
